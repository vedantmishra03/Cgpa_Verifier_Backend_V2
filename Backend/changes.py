# from flask import Flask, request, jsonify
# from flask_cors import CORS
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter
# import cloudinary
# import cloudinary.uploader
# from io import BytesIO
# import time
# import logging

# logging.basicConfig(level=logging.DEBUG)
# app = Flask(__name__)

# # Enable CORS with specific settings
# CORS(app, resources={
#     r"/*": {
#         "origins": "*",
#         "methods": ["GET", "POST", "OPTIONS"],
#         "allow_headers": ["Content-Type"]
#     }
# })

# # Configure Cloudinary
# cloudinary.config(
#     cloud_name='diqxosnks',
#     api_key='445594711426492',
#     api_secret='RHoIZSUby9fKdXBXV7LcqhmvYAU'
# )

# # Path to the stored master sheet
# MASTER_SHEET_PATH = "https://res.cloudinary.com/diqxosnks/raw/upload/v1735944444/master_sheet_l8fklq.xlsx"

# @app.route('/')
# def home():
#     return "Welcome to the Excel Verifier Backend API. Use the `/process` endpoint to upload the response sheet."

# @app.route('/process', methods=['POST'])
# def process_file():
#     if 'response_sheet' not in request.files:
#         return {"error": "response_sheet.xlsx is required!"}, 400

#     response_file = request.files['response_sheet']

#     # Load the response sheet
#     try:
#         response_sheet = pd.read_excel(response_file)
#     except Exception as e:
#         logging.error(f"Excel read error: {str(e)}")
#         return {"error": "Invalid file format. Please upload a valid Excel file."}, 400

#     # Normalize column names
#     response_sheet.columns = response_sheet.columns.str.strip().str.lower().str.replace(' ', '_')
#     logging.debug(f"Normalized response sheet columns: {response_sheet.columns.tolist()}")

#     # Identify Rollno and Cgpa columns
#     rollno_column = next((col for col in response_sheet.columns if 'roll' in col), None)
#     cgpa_column = next((col for col in response_sheet.columns if 'cgpa' in col), None)
#     logging.debug(f"Identified rollno column: {rollno_column}, cgpa column: {cgpa_column}")

#     if not rollno_column or not cgpa_column:
#         return {"error": "response_sheet must contain 'Rollno' and 'Cgpa' columns."}, 400

#     try:
#         # Load the master sheet
#         master_sheet = pd.read_excel(MASTER_SHEET_PATH)

#         # Normalize column names in master sheet
#         master_sheet.columns = master_sheet.columns.str.strip().str.lower().str.replace(' ', '_')
#         logging.debug(f"Normalized master sheet columns: {master_sheet.columns.tolist()}")

#         # Merge with left join to keep all response entries
#         merged_data = response_sheet.merge(master_sheet, left_on=rollno_column, right_on='rollno', how="left", suffixes=('_response', '_master'))
#         logging.debug(f"Merged data columns: {merged_data.columns.tolist()}")

#         # Check for invalid roll numbers and updated CGPA
#         merged_data['invalid'] = merged_data['cgpa_master'].isna()
#         merged_data['changed'] = (merged_data[f"{cgpa_column}_response"] != merged_data['cgpa_master']) & ~merged_data['invalid']

#         # Prepare the output data
#         merged_data['correct_cgpa'] = merged_data.apply(
#             lambda row: row['cgpa_master'] if not row['invalid'] and row['changed'] else row[f"{cgpa_column}_response"],
#             axis=1
#         )

#         # Insert the Correct_Cgpa column next to the identified Cgpa column
#         cgpa_index = merged_data.columns.get_loc(f"{cgpa_column}_response") + 1
#         merged_data.insert(cgpa_index, 'correct_cgpa', merged_data.pop('correct_cgpa'))

#         # Remove unwanted columns
#         columns_to_remove = ['column1', 'rollno', 'cgpa_master', 'invalid', 'changed']
#         merged_data = merged_data.drop(columns=columns_to_remove, errors='ignore')
#         logging.debug(f"Final data columns: {merged_data.columns.tolist()}")

#         # Save updated data to an in-memory file
#         output = BytesIO()
#         merged_data.to_excel(output, index=False, engine='openpyxl')

#         # Load and highlight in Excel
#         output.seek(0)
#         workbook = load_workbook(output)
#         worksheet = workbook.active

#         # Define cell fills
#         green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
#         red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

#         # Get column indices for the final Excel file
#         columns = list(merged_data.columns)
#         rollno_idx = columns.index(rollno_column) + 1
#         cgpa_idx = columns.index(f"{cgpa_column}_response") + 1  
#         correct_cgpa_idx = columns.index('correct_cgpa') + 1

#         # Convert column numbers to Excel letters
#         rollno_col = get_column_letter(rollno_idx)
#         cgpa_col = get_column_letter(cgpa_idx)
#         correct_cgpa_col = get_column_letter(correct_cgpa_idx)

#         logging.debug(f"Column letters: rollno={rollno_col}, cgpa={cgpa_col}, correct_cgpa={correct_cgpa_col}")

#         # Apply highlighting
#         for idx, row in merged_data.iterrows():
#             excel_row = idx + 2  # Excel rows start at 1, and we have header row
            
#             # Get cell references using column letters
#             rollno_cell = worksheet[f"{rollno_col}{excel_row}"]
#             cgpa_response_cell = worksheet[f"{cgpa_col}{excel_row}"]
#             correct_cgpa_cell = worksheet[f"{correct_cgpa_col}{excel_row}"]

#             # Mark invalid roll numbers
#             if row['invalid']:
#                 rollno_cell.fill = red_fill
            
#             # Mark incorrect CGPA values
#             if not row['invalid'] and row['changed']:
#                 cgpa_response_cell.fill = red_fill  # Original incorrect CGPA
#                 correct_cgpa_cell.fill = green_fill  # Corrected CGPA value

#         # Save to bytes buffer
#         output = BytesIO()
#         workbook.save(output)
#         output_bytes = output.getvalue()

#         # Upload bytes directly to Cloudinary with unique filename
#         try:
#             timestamp = int(time.time())
#             upload_result = cloudinary.uploader.upload(
#                 output_bytes,
#                 resource_type="auto",  # Ensure the file is recognized as an Excel file
#                 folder="excel_files",
#                 public_id=f"updated_response_sheet_{timestamp}",
#                 unique_filename=True,
#                 format="xlsx" 
#             )
            
#             if 'secure_url' not in upload_result:
#                 logging.error("Cloudinary upload failed - no secure_url in response")
#                 return {"error": "Failed to upload file to Cloudinary"}, 500
                
#             return jsonify({"download_url": upload_result['secure_url']})
            
#         except Exception as e:
#             logging.error(f"Cloudinary upload error: {str(e)}")
#             return {"error": "Failed to upload file to Cloudinary"}, 500

#     except Exception as e:
#         logging.error(f"Processing error: {str(e)}")
#         return {"error": f"Error processing the file: {str(e)}"}, 500





# #################################################################
#  initail code 

# from flask import Flask, request, jsonify
# from flask_cors import CORS
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import cloudinary
# import cloudinary.uploader
# from io import BytesIO
# import time
# import logging

# logging.basicConfig(level=logging.DEBUG)
# app = Flask(__name__)

# # Enable CORS with specific settings
# CORS(app, resources={
#     r"/*": {
#         "origins": "*",
#         "methods": ["GET", "POST", "OPTIONS"],
#         "allow_headers": ["Content-Type"]
#     }
# })

# # Configure Cloudinary
# cloudinary.config(
#     cloud_name='diqxosnks',
#     api_key='445594711426492',
#     api_secret='RHoIZSUby9fKdXBXV7LcqhmvYAU'
# )

# # Path to the stored master sheet
# MASTER_SHEET_PATH = "https://res.cloudinary.com/diqxosnks/raw/upload/v1735944444/master_sheet_l8fklq.xlsx"

# @app.route('/')
# def home():
#     return "Welcome to the Excel Verifier Backend API. Use the `/process` endpoint to upload the response sheet."

# @app.route('/process', methods=['POST'])
# def process_file():
#     if 'response_sheet' not in request.files:
#         return {"error": "response_sheet.xlsx is required!"}, 400

#     response_file = request.files['response_sheet']

#     # Load the response sheet
#     try:
#         response_sheet = pd.read_excel(response_file)
#     except Exception as e:
#         logging.error(f"Excel read error: {str(e)}")
#         return {"error": "Invalid file format. Please upload a valid Excel file."}, 400

#     # Validate the format of the response sheet
#     required_columns = ["Rollno", "Cgpa"]
#     if not all(column in response_sheet.columns for column in required_columns):
#         return {"error": "response_sheet must be in the desired format with 'Rollno' and 'Cgpa' as the first two columns."}, 400

#     try:
#         # Load the master sheet
#         master_sheet = pd.read_excel(MASTER_SHEET_PATH)

#         # Merge with left join to keep all response entries
#         merged_data = response_sheet.merge(master_sheet, on="Rollno", how="left", suffixes=('_response', '_master'))

#         # Check for invalid roll numbers and updated CGPA
#         merged_data['Invalid'] = merged_data['Cgpa_master'].isna()
#         merged_data['Changed'] = (merged_data['Cgpa_response'] != merged_data['Cgpa_master']) & ~merged_data['Invalid']

#         # Prepare the output data
#         merged_data['Correct_Cgpa'] = merged_data.apply(
#             lambda row: row['Cgpa_master'] if not row['Invalid'] and row['Changed'] else row['Cgpa_response'],
#             axis=1
#         )

#         # Save updated data to an in-memory file
#         output = BytesIO()
#         merged_data[['Rollno', 'Cgpa_response', 'Correct_Cgpa']].rename(
#             columns={'Cgpa_response': 'Old_Cgpa'}
#         ).to_excel(output, index=False, engine='openpyxl')

#         # Load and highlight in Excel
#         output.seek(0)
#         workbook = load_workbook(output)
#         worksheet = workbook.active

#         # Define cell fills
#         green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
#         red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

#         # Apply highlighting
#         for row in range(2, worksheet.max_row + 1):
#             if worksheet[f"A{row}"].value in merged_data[merged_data['Invalid']]['Rollno'].values:  # Invalid Rollno
#                 worksheet[f"A{row}"].fill = red_fill
#             if worksheet[f"B{row}"].value != worksheet[f"C{row}"].value:  # Changed Cgpa
#                 worksheet[f"B{row}"].fill = red_fill
#                 worksheet[f"C{row}"].fill = green_fill

#         # Save to bytes buffer
#         output = BytesIO()
#         workbook.save(output)
#         output_bytes = output.getvalue()

#         # Upload bytes directly to Cloudinary with unique filename
#         try:
#             timestamp = int(time.time())
#             upload_result = cloudinary.uploader.upload(
#                 output_bytes,
#                 resource_type="raw",
#                 folder="excel_files",
#                 public_id=f"updated_response_sheet_{timestamp}",
#                 # format="xlsx",
#                 unique_filename=True
#             )
            
#             if 'secure_url' not in upload_result:
#                 logging.error("Cloudinary upload failed - no secure_url in response")
#                 return {"error": "Failed to upload file to Cloudinary"}, 500
                
#             return jsonify({"download_url": upload_result['secure_url']})
            
#         except Exception as e:
#             logging.error(f"Cloudinary upload error: {str(e)}")
#             return {"error": "Failed to upload file to Cloudinary"}, 500

#     except Exception as e:
#         logging.error(f"Processing error: {str(e)}")
#         return {"error": "Error processing the file"}, 500

########################################################################Final coed 
# from flask import Flask, request, jsonify
# from flask_cors import CORS
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import cloudinary
# import cloudinary.uploader
# from io import BytesIO
# import time
# import logging
# from openpyxl.utils import get_column_letter

# logging.basicConfig(level=logging.DEBUG)
# app = Flask(__name__)

# # Enable CORS with specific settings
# CORS(app, resources={
#     r"/*": {
#         "origins": "*",
#         "methods": ["GET", "POST", "OPTIONS"],
#         "allow_headers": ["Content-Type"]
#     }
# })

# # Configure Cloudinary
# cloudinary.config(
#     cloud_name='diqxosnks',
#     api_key='445594711426492',
#     api_secret='RHoIZSUby9fKdXBXV7LcqhmvYAU'
# )

# # Path to the stored master sheet
# MASTER_SHEET_PATH = "https://res.cloudinary.com/diqxosnks/raw/upload/v1735944444/master_sheet_l8fklq.xlsx"

# @app.route('/')
# def home():
#     return "Welcome to the Excel Verifier Backend API. Use the `/process` endpoint to upload the response sheet."

# @app.route('/process', methods=['POST'])
# def process_file():
#     if 'response_sheet' not in request.files:
#         return {"error": "response_sheet.xlsx is required!"}, 400

#     response_file = request.files['response_sheet']

#     # Load the response sheet
#     try:
#         response_sheet = pd.read_excel(response_file)
#     except Exception as e:
#         logging.error(f"Excel read error: {str(e)}")
#         return {"error": "Invalid file format. Please upload a valid Excel file."}, 400

#     # Normalize column names
#     response_sheet.columns = response_sheet.columns.str.strip().str.lower().str.replace(' ', '_')
#     logging.debug(f"Normalized response sheet columns: {response_sheet.columns.tolist()}")

#     # Identify Rollno and Cgpa columns
#     rollno_column = next((col for col in response_sheet.columns if 'roll' in col), None)
#     cgpa_column = next((col for col in response_sheet.columns if 'cgpa' in col), None)
#     logging.debug(f"Identified rollno column: {rollno_column}, cgpa column: {cgpa_column}")

#     if not rollno_column or not cgpa_column:
#         return {"error": "response_sheet must contain 'Rollno' and 'Cgpa' columns."}, 400

#     try:
#         # Load the master sheet
#         master_sheet = pd.read_excel(MASTER_SHEET_PATH)

#         # Normalize column names in master sheet
#         master_sheet.columns = master_sheet.columns.str.strip().str.lower().str.replace(' ', '_')
#         logging.debug(f"Normalized master sheet columns: {master_sheet.columns.tolist()}")

#         # Merge with left join to keep all response entries
#         merged_data = response_sheet.merge(master_sheet, left_on=rollno_column, right_on='rollno', how="left", suffixes=('_response', '_master'))
#         logging.debug(f"Merged data columns: {merged_data.columns.tolist()}")

#         # Check for invalid roll numbers and updated CGPA
#         merged_data['invalid'] = merged_data['cgpa_master'].isna()
#         merged_data['changed'] = (merged_data[f"{cgpa_column}_response"] != merged_data['cgpa_master']) & ~merged_data['invalid']

#         # Prepare the output data
#         merged_data['correct_cgpa'] = merged_data.apply(
#             lambda row: row['cgpa_master'] if not row['invalid'] and row['changed'] else row[f"{cgpa_column}_response"],
#             axis=1
#         )

#         # Insert the Correct_Cgpa column next to the identified Cgpa column
#         cgpa_index = merged_data.columns.get_loc(f"{cgpa_column}_response") + 1
#         merged_data.insert(cgpa_index, 'correct_cgpa', merged_data.pop('correct_cgpa'))

#         # Save updated data to an in-memory file
#         output = BytesIO()
#         merged_data.to_excel(output, index=False, engine='openpyxl')

#         # Load and highlight in Excel
#         output.seek(0)
#         workbook = load_workbook(output)
#         worksheet = workbook.active

#         # Define cell fills
#         green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
#         red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

#         # Get column indices for the final Excel file
#         columns = list(merged_data.columns)
#         rollno_idx = columns.index(rollno_column) + 1
#         cgpa_idx = columns.index(f"{cgpa_column}_response") + 1  
#         correct_cgpa_idx = columns.index('correct_cgpa') + 1

#         # Convert column numbers to Excel letters
#         rollno_col = get_column_letter(rollno_idx)
#         cgpa_col = get_column_letter(cgpa_idx)
#         correct_cgpa_col = get_column_letter(correct_cgpa_idx)

#         logging.debug(f"Column letters: rollno={rollno_col}, cgpa={cgpa_col}, correct_cgpa={correct_cgpa_col}")

#         # Apply highlighting
#         for idx, row in merged_data.iterrows():
#             excel_row = idx + 2  # Excel rows start at 1, and we have header row
            
#             # Get cell references using column letters
#             rollno_cell = worksheet[f"{rollno_col}{excel_row}"]
#             cgpa_response_cell = worksheet[f"{cgpa_col}{excel_row}"]
#             correct_cgpa_cell = worksheet[f"{correct_cgpa_col}{excel_row}"]

#             # Mark invalid roll numbers
#             if row['invalid']:
#                 rollno_cell.fill = red_fill
            
#             # Mark incorrect CGPA values
#             if not row['invalid'] and row['changed']:
#                 cgpa_response_cell.fill = red_fill  # Original incorrect CGPA
#                 correct_cgpa_cell.fill = green_fill  # Corrected CGPA value

#         # Save to bytes buffer
#         output = BytesIO()
#         workbook.save(output)
#         output_bytes = output.getvalue()

#         # Upload bytes directly to Cloudinary with unique filename
#         try:
#             timestamp = int(time.time())
#             upload_result = cloudinary.uploader.upload(
#                output_bytes,
#                 resource_type="raw",
#                 folder="excel_files",
#                 public_id=f"updated_response_sheet_{timestamp}",
#                 format="xlsx",
#                 unique_filename=True
#             )
            
#             if 'secure_url' not in upload_result:
#                 logging.error("Cloudinary upload failed - no secure_url in response")
#                 return {"error": "Failed to upload file to Cloudinary"}, 500
                
#             return jsonify({"download_url": upload_result['secure_url']})
            
#         except Exception as e:
#             logging.error(f"Cloudinary upload error: {str(e)}")
#             return {"error": "Failed to upload file to Cloudinary"}, 500

#     except Exception as e:
#         logging.error(f"Processing error: {str(e)}")
#         return {"error": f"Error processing the file: {str(e)}"}, 500



# //////using the master sheet in sted of the cloudinamry 
from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import cloudinary
import cloudinary.uploader
from io import BytesIO
import time
import logging
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()

logging.basicConfig(level=logging.DEBUG)
app = Flask(__name__)

# Enable CORS with specific settings
CORS(app, resources={
    r"/*": {
        "origins": "*",
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# Configure Cloudinary
cloudinary.config(
    cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME'),
    api_key=os.getenv('CLOUDINARY_API_KEY'),
    api_secret=os.getenv('CLOUDINARY_API_SECRET')
)

# Path to the stored master sheet
MASTER_SHEET_PATH = os.getenv('MASTER_SHEET_PATH')

@app.route('/')
def home():
    return "Welcome to the Excel Verifier Backend API. Use the `/process` endpoint to upload the response sheet."

@app.route('/process', methods=['POST'])
def process_file():
    if 'response_sheet' not in request.files:
        return {"error": "response_sheet.xlsx is required!"}, 400

    response_file = request.files['response_sheet']

    # Load the response sheet
    try:
        response_sheet = pd.read_excel(response_file, index_col=None)  # Added index_col=None to prevent 'column1'
    except Exception as e:
        logging.error(f"Excel read error: {str(e)}")
        return {"error": "Invalid file format. Please upload a valid Excel file."}, 400

    # Normalize column names
    response_sheet.columns = response_sheet.columns.str.strip().str.lower().str.replace(' ', '_')
    logging.debug(f"Normalized response sheet columns: {response_sheet.columns.tolist()}")

    # Identify Rollno and Cgpa columns
    rollno_column = next((col for col in response_sheet.columns if 'roll' in col), None)
    cgpa_column = next((col for col in response_sheet.columns if 'cgpa' in col), None)
    logging.debug(f"Identified rollno column: {rollno_column}, cgpa column: {cgpa_column}")

    if not rollno_column or not cgpa_column:
        return {"error": "response_sheet must contain 'Rollno' and 'Cgpa' columns."}, 400

    try:
        # Load the master sheet with explicit error handling
        try:
            master_sheet = pd.read_csv(MASTER_SHEET_PATH, encoding='utf-8')
            logging.debug(f"Successfully loaded master sheet from {MASTER_SHEET_PATH}")
        except UnicodeDecodeError as e:
            logging.error(f"Unicode decode error: {str(e)}")
            return {"error": "Failed to decode master data. Please check the file encoding."}, 500
        except Exception as csv_error:
            logging.error(f"Failed to load CSV: {str(csv_error)}")
            return {"error": "Failed to load master data"}, 500

        # Rest remains same
        master_sheet.columns = master_sheet.columns.str.strip().str.lower().str.replace(' ', '_')
        logging.debug(f"Normalized master sheet columns: {master_sheet.columns.tolist()}")

        merged_data = response_sheet.merge(
            master_sheet, 
            left_on=rollno_column, 
            right_on='rollno', 
            how="left", 
            suffixes=('_response', '_master')
        )
        logging.debug(f"Merged data columns: {merged_data.columns.tolist()}")

        # Check for invalid roll numbers and updated CGPA
        invalid_rolls = merged_data['cgpa_master'].isna()
        changed_cgpa = (merged_data[f"{cgpa_column}_response"] != merged_data['cgpa_master']) & ~invalid_rolls

        # Prepare the output data with correct CGPA
        merged_data['correct_cgpa'] = merged_data.apply(
            lambda row: row['cgpa_master'] if not pd.isna(row['cgpa_master']) and row[f"{cgpa_column}_response"] != row['cgpa_master']
            else row[f"{cgpa_column}_response"],
            axis=1
        )

        # Create final dataframe with only required columns
        final_data = response_sheet.copy()
        
        # Insert the Correct_Cgpa column next to the original Cgpa column
        cgpa_idx = list(final_data.columns).index(cgpa_column)
        final_data.insert(cgpa_idx + 1, 'correct_cgpa', merged_data['correct_cgpa'])

        # Save updated data to an in-memory file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            final_data.to_excel(writer, index=False, sheet_name='Sheet1')

        # Load and highlight in Excel
        output.seek(0)
        workbook = load_workbook(output)
        worksheet = workbook.active

        # Define cell fills
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Get column indices for the final Excel file
        columns = list(final_data.columns)
        rollno_idx = columns.index(rollno_column) + 1
        cgpa_idx = columns.index(cgpa_column) + 1
        correct_cgpa_idx = columns.index('correct_cgpa') + 1

        # Convert column numbers to Excel letters
        rollno_col = get_column_letter(rollno_idx)
        cgpa_col = get_column_letter(cgpa_idx)
        correct_cgpa_col = get_column_letter(correct_cgpa_idx)

        logging.debug(f"Column letters: rollno={rollno_col}, cgpa={cgpa_col}, correct_cgpa={correct_cgpa_col}")

        # Update the highlighting section
        for idx, (_, row) in enumerate(merged_data.iterrows(), start=2):
            # Get cell references
            rollno_cell = worksheet[f"{rollno_col}{idx}"]
            cgpa_response_cell = worksheet[f"{cgpa_col}{idx}"]
            correct_cgpa_cell = worksheet[f"{correct_cgpa_col}{idx}"]

            # Clear all fills first
            rollno_cell.fill = PatternFill()
            cgpa_response_cell.fill = PatternFill()
            correct_cgpa_cell.fill = PatternFill()

            # Case 1: Roll number not in master sheet
            if pd.isna(row['cgpa_master']):
                rollno_cell.fill = red_fill
            
            # Case 2: Roll number exists in master sheet
            else:
                # Check if CGPA values are different
                if row[f"{cgpa_column}_response"] != row['cgpa_master']:
                    cgpa_response_cell.fill = red_fill  # Wrong CGPA
                    correct_cgpa_cell.fill = green_fill  # Correct CGPA

        # Save to bytes buffer
        output = BytesIO()
        workbook.save(output)
        output_bytes = output.getvalue()

        # Upload bytes directly to Cloudinary with unique filename
        try:
            timestamp = int(time.time())
            upload_result = cloudinary.uploader.upload(
                output_bytes,
                resource_type="raw",
                folder="excel_files",
                public_id=f"updated_response_sheet_{timestamp}",
                format="xlsx",
                unique_filename=True
            )
            
            if 'secure_url' not in upload_result:
                logging.error("Cloudinary upload failed - no secure_url in response")
                return {"error": "Failed to upload file to Cloudinary"}, 500
                
            return jsonify({"download_url": upload_result['secure_url']})
            
        except Exception as e:
            logging.error(f"Cloudinary upload error: {str(e)}")
            return {"error": "Failed to upload file to Cloudinary"}, 500

    except Exception as e:
        logging.error(f"Processing error: {str(e)}")
        return {"error": f"Error processing the file: {str(e)}"}, 500